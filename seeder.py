#!/usr/bin/env python3
"""
seeder.py — Idempotent seeder script for DSS Sports Inter-House Meet.
Parses /seed-data/interhouse_meet.xlsx and seeds Supabase DB.
"""

import os
import re
import sys
import uuid
import argparse
from typing import Dict, Any, Optional, Tuple, List
import openpyxl
from dotenv import load_dotenv

load_dotenv()

SUPABASE_URL = os.getenv("SUPABASE_URL")
SUPABASE_SERVICE_ROLE_KEY = os.getenv("SUPABASE_SERVICE_ROLE_KEY") or os.getenv("SUPABASE_KEY")

HOUSE_CONFIG = [
    {"name": "Karnali", "color_hex": "#10B981", "short_code": "KAR"},
    {"name": "Koshi", "color_hex": "#0EA5E9", "short_code": "KOS"},
    {"name": "Mahakali", "color_hex": "#8B5CF6", "short_code": "MAH"},
    {"name": "Mechi", "color_hex": "#F97316", "short_code": "MEC"}
]

SPORT_CONFIG = [
    {"name": "Futsal", "type": "football", "level": "HS"},
    {"name": "Basketball", "type": "basketball", "level": "HS"},
    {"name": "Cricksal", "type": "generic", "level": "HS"}
]

TOURNAMENT_GROUP_CONFIG = [
    {"sport_name": "Futsal", "gender": "Boys", "format": "pool_to_semis", "point_win": 3, "point_draw": 1, "point_loss": 0},
    {"sport_name": "Futsal", "gender": "Girls", "format": "round_robin", "point_win": 3, "point_draw": 1, "point_loss": 0},
    {"sport_name": "Basketball", "gender": "Boys", "format": "round_robin", "point_win": 3, "point_draw": 1, "point_loss": 0},
    {"sport_name": "Basketball", "gender": "Girls", "format": "round_robin", "point_win": 3, "point_draw": 1, "point_loss": 0},
    {"sport_name": "Cricksal", "gender": "Boys", "format": "pool_to_semis", "point_win": 3, "point_draw": 1, "point_loss": 0},
    {"sport_name": "Cricksal", "gender": "Girls", "format": "round_robin", "point_win": 3, "point_draw": 1, "point_loss": 0}
]

# Squad count matrix: {sport: {house: {gender: count}}}
SQUAD_COUNT_MATRIX = {
    "Cricksal": {
        "Karnali": {"Boys": 2, "Girls": 1},
        "Koshi": {"Boys": 2, "Girls": 1},
        "Mahakali": {"Boys": 1, "Girls": 1},
        "Mechi": {"Boys": 2, "Girls": 1}
    },
    "Futsal": {
        "Karnali": {"Boys": 2, "Girls": 2},
        "Koshi": {"Boys": 2, "Girls": 1},
        "Mahakali": {"Boys": 1, "Girls": 1},
        "Mechi": {"Boys": 2, "Girls": 1}
    },
    "Basketball": {
        "Karnali": {"Boys": 2, "Girls": 2},
        "Koshi": {"Boys": 1, "Girls": 1},
        "Mahakali": {"Boys": 1, "Girls": 1},
        "Mechi": {"Boys": 1, "Girls": 1}
    }
}

class InterHouseSeeder:
    def __init__(self, excel_path: str, supabase_client=None):
        self.excel_path = excel_path
        self.client = supabase_client
        self.wb = None

        self.houses_map: Dict[str, Dict[str, Any]] = {}
        self.sports_map: Dict[str, Dict[str, Any]] = {}
        self.groups_map: Dict[str, Dict[str, Any]] = {}
        self.squads_map: Dict[str, Dict[str, Any]] = {}
        self.players_map: Dict[str, Dict[str, Any]] = {}

        self.stats = {
            "houses": {"created": 0, "updated": 0, "skipped": 0},
            "sports": {"created": 0, "updated": 0, "skipped": 0},
            "groups": {"created": 0, "updated": 0, "skipped": 0},
            "squads": {"created": 0, "updated": 0, "skipped": 0},
            "players": {"created": 0, "updated": 0, "skipped": 0},
            "fixtures": {
                "created": 0,
                "updated": 0,
                "unplayed": 0,
                "unparseable": 0,
                "unparseable_details": []
            }
        }

    def load_workbook(self) -> bool:
        if not os.path.exists(self.excel_path):
            print(f"Warning: Excel file not found at '{self.excel_path}'. Proceeding with default config.")
            return False
        self.wb = openpyxl.load_workbook(self.excel_path, data_only=True)
        return True

    def seed_houses(self):
        print("\n=== Seeding Houses ===")
        for house in HOUSE_CONFIG:
            key = house["name"].lower()
            house_uuid = str(uuid.uuid5(uuid.NAMESPACE_DNS, f"house.{house['name']}"))
            record = {
                "id": house_uuid,
                "name": house["name"],
                "color_hex": house["color_hex"],
                "short_code": house["short_code"]
            }

            if self.client:
                try:
                    res = self.client.table("houses").upsert(
                        record, on_conflict="name"
                    ).execute()
                    if res.data:
                        record = res.data[0]
                    self.stats["houses"]["created"] += 1
                except Exception as e:
                    print(f"Error upserting house {house['name']}: {e}")
                    self.stats["houses"]["skipped"] += 1
            else:
                self.stats["houses"]["created"] += 1

            self.houses_map[key] = record
            print(f"  House: {record['name']} ({record['short_code']}) -> {record['color_hex']}")

    def seed_sports(self):
        print("\n=== Seeding Sports & Tournament Groups ===")
        for sport in SPORT_CONFIG:
            key = sport["name"].lower()
            sport_uuid = str(uuid.uuid5(uuid.NAMESPACE_DNS, f"sport.{sport['name']}"))
            record = {
                "id": sport_uuid,
                "name": sport["name"],
                "type": sport["type"],
                "level": sport["level"]
            }

            if self.client:
                try:
                    res = self.client.table("sports").upsert(
                        record, on_conflict="name"
                    ).execute()
                    if res.data:
                        record = res.data[0]
                    self.stats["sports"]["created"] += 1
                except Exception as e:
                    print(f"Error upserting sport {sport['name']}: {e}")
                    self.stats["sports"]["skipped"] += 1
            else:
                self.stats["sports"]["created"] += 1

            self.sports_map[key] = record
            print(f"  Sport: {record['name']} ({record['type']})")

        # Seed Tournament Groups
        for grp in TOURNAMENT_GROUP_CONFIG:
            sport_obj = self.sports_map.get(grp["sport_name"].lower())
            if not sport_obj:
                continue
            grp_key = f"{sport_obj['id']}_{grp['gender']}"
            grp_uuid = str(uuid.uuid5(uuid.NAMESPACE_DNS, f"group.{sport_obj['name']}.{grp['gender']}"))
            grp_record = {
                "id": grp_uuid,
                "sport_id": sport_obj["id"],
                "gender": grp["gender"],
                "format": grp["format"],
                "point_win": grp["point_win"],
                "point_draw": grp["point_draw"],
                "point_loss": grp["point_loss"]
            }

            if self.client:
                try:
                    res = self.client.table("tournament_groups").upsert(
                        grp_record, on_conflict="sport_id,gender"
                    ).execute()
                    if res.data:
                        grp_record = res.data[0]
                    self.stats["groups"]["created"] += 1
                except Exception as e:
                    print(f"Error upserting tournament group {grp['sport_name']} {grp['gender']}: {e}")
                    self.stats["groups"]["skipped"] += 1
            else:
                self.stats["groups"]["created"] += 1

            self.groups_map[grp_key] = grp_record
            print(f"  Tournament Group: {grp['sport_name']} {grp['gender']} -> {grp['format']}")

    def ensure_default_squads(self):
        for h_key, house_obj in self.houses_map.items():
            h_name = house_obj["name"]
            for s_key, sport_obj in self.sports_map.items():
                s_name = sport_obj["name"]
                for gender in ["Boys", "Girls"]:
                    count = SQUAD_COUNT_MATRIX.get(s_name, {}).get(h_name, {}).get(gender, 1)
                    labels = ["A", "B"] if count == 2 else ["Single"]
                    for squad_label in labels:
                        squad_key = f"{house_obj['id']}_{sport_obj['id']}_{gender}_{squad_label}"
                        if squad_key not in self.squads_map:
                            if squad_label == "Single":
                                squad_name = f"{house_obj['name']}"
                            else:
                                squad_name = f"{house_obj['name']} {squad_label}"
                            sq_uuid = str(uuid.uuid5(uuid.NAMESPACE_DNS, f"squad.{house_obj['name']}.{sport_obj['name']}.{gender}.{squad_label}"))
                            squad_record = {
                                "id": sq_uuid,
                                "name": squad_name,
                                "house_id": house_obj["id"],
                                "gender": gender,
                                "squad_label": squad_label,
                                "pool": None,
                                "sport_id": sport_obj["id"],
                                "level": "HS"
                            }
                            if self.client:
                                try:
                                    res = self.client.table("teams").upsert(
                                        squad_record, on_conflict="house_id,sport_id,gender,squad_label,level"
                                    ).execute()
                                    if res.data:
                                        squad_record = res.data[0]
                                    self.stats["squads"]["created"] += 1
                                except Exception:
                                    self.stats["squads"]["skipped"] += 1
                            else:
                                self.stats["squads"]["created"] += 1

                            self.squads_map[squad_key] = squad_record

    def parse_house_rosters(self):
        if not self.wb:
            return

        print("\n=== Parsing House Rosters (Squads & Players) ===")
        house_sheets = [s for s in self.wb.sheetnames if "house" in s.lower() and s != "House Members"]

        for sheet_name in house_sheets:
            sheet = self.wb[sheet_name]
            house_name = sheet_name.replace("House", "").strip()
            house_key = house_name.lower()
            house_obj = self.houses_map.get(house_key)

            if not house_obj:
                print(f"  Skipping sheet '{sheet_name}': House '{house_name}' not in house map")
                continue

            print(f"  Processing sheet: {sheet_name} (House: {house_obj['name']})")
            max_cols = sheet.max_column

            block_starts = []
            for col in range(1, max_cols + 1):
                val = sheet.cell(row=1, column=col).value
                if val and isinstance(val, str):
                    for s_key in self.sports_map.keys():
                        if s_key in val.lower():
                            block_starts.append((col, self.sports_map[s_key]))
                            break

            if not block_starts:
                for col in range(1, max_cols + 1):
                    val = sheet.cell(row=2, column=col).value
                    if val and isinstance(val, str):
                        for s_key in self.sports_map.keys():
                            if s_key in val.lower():
                                block_starts.append((col, self.sports_map[s_key]))
                                break

            for start_col, sport_obj in block_starts:
                header_row = 2
                col_indices = {}
                for c in range(start_col, min(start_col + 10, max_cols + 1)):
                    cell_val = sheet.cell(row=header_row, column=c).value
                    if not cell_val:
                        cell_val = sheet.cell(row=header_row + 1, column=c).value
                    if cell_val and isinstance(cell_val, str):
                        c_lower = cell_val.strip().lower()
                        if "roll" in c_lower or "rn" in c_lower:
                            col_indices["roll"] = c
                        elif "name" in c_lower:
                            col_indices["name"] = c
                        elif "grade" in c_lower:
                            col_indices["grade"] = c
                        elif "sec" in c_lower:
                            col_indices["section"] = c
                        elif "gen" in c_lower or "sex" in c_lower:
                            col_indices["gender"] = c
                        elif "team" in c_lower or "squad" in c_lower:
                            col_indices["squad"] = c

                data_start_row = 3 if "name" in [str(sheet.cell(row=2, column=c).value).lower() for c in range(start_col, start_col + 8)] else 4

                for r in range(data_start_row, sheet.max_row + 1):
                    name_col = col_indices.get("name", start_col + 2)
                    name_val = sheet.cell(row=r, column=name_col).value
                    if not name_val or not str(name_val).strip() or str(name_val).strip().isdigit():
                        continue

                    p_name = str(name_val).strip()
                    roll_num = str(sheet.cell(row=r, column=col_indices.get("roll", start_col + 1)).value or "").strip()
                    grade = str(sheet.cell(row=r, column=col_indices.get("grade", start_col + 3)).value or "").strip()
                    section = str(sheet.cell(row=r, column=col_indices.get("section", start_col + 4)).value or "").strip()
                    gender_raw = str(sheet.cell(row=r, column=col_indices.get("gender", start_col + 5)).value or "").strip()
                    squad_label_raw = str(sheet.cell(row=r, column=col_indices.get("squad", start_col + 6)).value or "").strip().upper()

                    gender = "Girls" if "girl" in gender_raw.lower() or "f" in gender_raw.lower() else "Boys"
                    squad_label = squad_label_raw if squad_label_raw in ["A", "B", "C", "D"] else "A"

                    squad_key = f"{house_obj['id']}_{sport_obj['id']}_{gender}_{squad_label}"

                    if squad_key not in self.squads_map:
                        squad_name = f"{house_obj['name']} {gender} {sport_obj['name']} {squad_label}"
                        sq_uuid = str(uuid.uuid5(uuid.NAMESPACE_DNS, f"squad.{house_obj['name']}.{sport_obj['name']}.{gender}.{squad_label}"))
                        squad_record = {
                            "id": sq_uuid,
                            "name": squad_name,
                            "house_id": house_obj["id"],
                            "gender": gender,
                            "squad_label": squad_label,
                            "pool": None,
                            "sport_id": sport_obj["id"],
                            "level": "HS"
                        }
                        self.squads_map[squad_key] = squad_record
                        if self.client:
                            try:
                                res = self.client.table("teams").upsert(
                                    squad_record, on_conflict="house_id,sport_id,gender,squad_label,level"
                                ).execute()
                                if res.data:
                                    squad_record = res.data[0]
                                self.stats["squads"]["created"] += 1
                            except Exception:
                                self.stats["squads"]["skipped"] += 1
                        else:
                            self.stats["squads"]["created"] += 1

                        self.squads_map[squad_key] = squad_record

                    squad_obj = self.squads_map[squad_key]

                    p_key = roll_num if roll_num else f"{p_name}_{squad_obj['id']}"
                    player_uuid = str(uuid.uuid5(uuid.NAMESPACE_DNS, f"player.{roll_num if roll_num else p_key}"))
                    p_record = {
                        "id": player_uuid,
                        "name": p_name,
                        "team_id": squad_obj["id"],
                        "roll_number": roll_num if roll_num else None,
                        "grade": grade,
                        "section": section,
                        "gender": gender,
                        "level": "HS"
                    }

                    if self.client:
                        try:
                            res = self.client.table("players").upsert(
                                p_record, on_conflict="roll_number" if roll_num else "id"
                            ).execute()
                            if res.data:
                                p_record = res.data[0]
                            self.stats["players"]["created"] += 1
                        except Exception:
                            self.stats["players"]["skipped"] += 1
                    else:
                        self.stats["players"]["created"] += 1

                    self.players_map[p_key] = p_record

    def parse_fixtures_and_scores(self):
        if not self.wb or "Tie-sheet & Scores Update" not in self.wb.sheetnames:
            return

        print("\n=== Parsing Fixtures & Scores ('Tie-sheet & Scores Update') ===")
        sheet = self.wb["Tie-sheet & Scores Update"]

        current_sport_name = "Futsal"
        current_gender = "Boys"

        for r in range(1, sheet.max_row + 1):
            row_vals = [sheet.cell(row=r, column=c).value for c in range(1, sheet.max_column + 1)]
            row_str = " ".join([str(v) for v in row_vals if v is not None]).strip()
            if not row_str:
                continue

            row_lower = row_str.lower()

            # Update active sport/gender context if header row contains sport/gender keywords
            for s_key, s_obj in self.sports_map.items():
                if s_key in row_lower:
                    current_sport_name = s_obj["name"]
                    break

            if "girl" in row_lower and " vs " not in row_lower:
                current_gender = "Girls"
            elif "boy" in row_lower and " vs " not in row_lower:
                current_gender = "Boys"

            sport_obj = self.sports_map.get(current_sport_name.lower(), list(self.sports_map.values())[0])

            # Parse Pool Headers for Format B (e.g., "Pole A: Mechi A, Koshi B... Pool B: Koshi A...")
            if ("pole a" in row_lower or "pool a" in row_lower) and ("pole b" in row_lower or "pool b" in row_lower):
                self.parse_pool_header_row(row_str, sport_obj["id"], current_gender)
                continue

            # Parse Semi Final / Final Placeholders
            if "semi final" in row_lower or "semifinal" in row_lower:
                self.create_placeholder_match(sport_obj["id"], current_gender, "semifinal", "Semi Final Match", r)
                # If row mentions 2 matches
                if "2 match" in row_lower or "2 matches" in row_lower:
                    self.create_placeholder_match(sport_obj["id"], current_gender, "semifinal", "Semi Final Match", r)
                continue
            elif "final match" in row_lower or " final " in f" {row_lower} ":
                if "semi" not in row_lower:
                    self.create_placeholder_match(sport_obj["id"], current_gender, "final", "Final Match", r)
                    continue

            # Parse Format A (4-cell separated) or Format B (merged single cell "Vs")
            # First check for Format B merged fixture cells in row
            merged_vs_cells = []
            for val in row_vals:
                if val and isinstance(val, str) and re.search(r'\bvs\b', val, re.IGNORECASE):
                    merged_vs_cells.append(str(val).strip())

            if len(merged_vs_cells) >= 1 and not (len(row_vals) >= 4 and any(str(v).strip().lower() in ["vs", "vs."] for v in row_vals if v)):
                # Format B: Process each merged cell in row
                for fixture_text in merged_vs_cells:
                    parts = re.split(r'\bvs\b', fixture_text, flags=re.IGNORECASE)
                    if len(parts) == 2:
                        team_a_str = parts[0].strip()
                        team_b_str = parts[1].strip()
                        self.process_fixture_pair(team_a_str, team_b_str, "", sport_obj, current_gender, r)
                continue

            # Format A: 4-cell separated (Team A | Vs | Team B | Score)
            if " vs " in row_lower or " vs. " in row_lower:
                team_a_str = ""
                team_b_str = ""
                score_str = ""

                for idx, val in enumerate(row_vals):
                    if val is None:
                        continue
                    v_str = str(val).strip()
                    if "vs" in v_str.lower():
                        if idx > 0 and row_vals[idx - 1]:
                            team_a_str = str(row_vals[idx - 1]).strip()
                        if idx + 1 < len(row_vals) and row_vals[idx + 1]:
                            team_b_str = str(row_vals[idx + 1]).strip()
                    if "=" in v_str or "-" in v_str:
                        if re.search(r'\d+\s*[-–]\s*\d+', v_str):
                            score_str = v_str

                if team_a_str and team_b_str:
                    self.process_fixture_pair(team_a_str, team_b_str, score_str, sport_obj, current_gender, r)

    def parse_pool_header_row(self, row_str: str, sport_id: str, gender: str):
        # Extract Pole A / Pool A and Pole B / Pool B contents
        pole_a_match = re.search(r'(?:Pole|Pool)\s*A[:\s]+(.*?)(?=(?:Pole|Pool)\s*B|$)', row_str, re.IGNORECASE)
        pole_b_match = re.search(r'(?:Pole|Pool)\s*B[:\s]+(.*)', row_str, re.IGNORECASE)

        if pole_a_match:
            teams_a_text = pole_a_match.group(1)
            for squad in self.squads_map.values():
                if squad["sport_id"] == sport_id and squad["gender"] == gender:
                    if self.is_squad_mentioned_in_text(squad, teams_a_text):
                        squad["pool"] = "A"
                        self.update_squad_pool(squad["id"], "A")

        if pole_b_match:
            teams_b_text = pole_b_match.group(1)
            for squad in self.squads_map.values():
                if squad["sport_id"] == sport_id and squad["gender"] == gender:
                    if self.is_squad_mentioned_in_text(squad, teams_b_text):
                        squad["pool"] = "B"
                        self.update_squad_pool(squad["id"], "B")

    def is_squad_mentioned_in_text(self, squad: Dict[str, Any], text: str) -> bool:
        h_name = self.houses_map.get(squad.get("house_id", ""), {}).get("name", "")
        if not h_name:
            for h in self.houses_map.values():
                if h["id"] == squad.get("house_id"):
                    h_name = h["name"]
                    break
        if not h_name:
            return False

        label = squad.get("squad_label", "Single")
        # Matches e.g. "Koshi B", "Mahakali", "Karnali A"
        pattern = r'\b' + re.escape(h_name) + (r'\s+' + label if label != "Single" else r'(\s+A|\s+B)?') + r'\b'
        return bool(re.search(pattern, text, re.IGNORECASE))

    def update_squad_pool(self, squad_id: str, pool_label: str):
        if self.client:
            try:
                self.client.table("teams").update({"pool": pool_label}).eq("id", squad_id).execute()
            except Exception as e:
                print(f"Error updating squad pool: {e}")

    def create_placeholder_match(self, sport_id: str, gender: str, stage: str, round_info: str, row_idx: int):
        match_uuid = str(uuid.uuid5(uuid.NAMESPACE_DNS, f"match.{sport_id}.{gender}.{stage}.placeholder.{row_idx}.{len(self.stats['fixtures'].get('created_list', []))}"))
        match_record = {
            "id": match_uuid,
            "sport_id": sport_id,
            "team_a_id": None,
            "team_b_id": None,
            "gender": gender,
            "stage": stage,
            "level": "HS",
            "status": "scheduled",
            "round_info": round_info,
            "winner_team_id": None,
            "is_draw": False,
            "score_team_a": 0,
            "score_team_b": 0,
            "score_difference": 0,
            "score_summary": ""
        }
        self.stats["fixtures"].setdefault("created_list", []).append(match_record)
        self.stats["fixtures"]["unplayed"] += 1

        if self.client:
            try:
                self.client.table("matches").insert(match_record).execute()
                self.stats["fixtures"]["created"] += 1
            except Exception as e:
                print(f"    Error inserting placeholder match: {e}")
        else:
            self.stats["fixtures"]["created"] += 1

    def process_fixture_pair(self, team_a_str: str, team_b_str: str, score_str: str, sport_obj: Dict[str, Any], gender: str, row_idx: int):
        squad_a = self.resolve_squad(team_a_str, sport_obj["id"], gender)
        squad_b = self.resolve_squad(team_b_str, sport_obj["id"], gender)

        if not squad_a or not squad_b:
            self.stats["fixtures"]["unparseable"] += 1
            self.stats["fixtures"]["unparseable_details"].append(
                f"Row {row_idx}: Could not resolve squad for '{team_a_str}' or '{team_b_str}'"
            )
            return

        status = "scheduled"
        winner_id = None
        is_draw = False
        score_a = 0
        score_b = 0
        score_diff = 0
        summary = ""

        if not score_str:
            self.stats["fixtures"]["unplayed"] += 1
        else:
            # Parse score: format X-Y=Z
            match_score = re.search(r'(\d+)\s*[-–]\s*(\d+)(?:\s*=\s*(\d+))?', score_str)
            if match_score:
                score_a = int(match_score.group(1))
                score_b = int(match_score.group(2))
                diff_calc = abs(score_a - score_b)

                # Arithmetic Verification Check if Z (printed diff) is present
                if match_score.group(3) is not None:
                    printed_diff = int(match_score.group(3))
                    if diff_calc != printed_diff:
                        self.stats["fixtures"]["unparseable"] += 1
                        self.stats["fixtures"]["unparseable_details"].append(
                            f"Row {row_idx}: Arithmetic mismatch |{score_a} - {score_b}| != printed difference {printed_diff} for match {squad_a['name']} vs {squad_b['name']}"
                        )
                        return

                score_diff = diff_calc
                status = "completed"
                summary = f"{score_a} - {score_b}"

                if score_a > score_b:
                    winner_id = squad_a["id"]
                elif score_b > score_a:
                    winner_id = squad_b["id"]
                else:
                    is_draw = True
            else:
                status = "scheduled"
                self.stats["fixtures"]["unparseable"] += 1
                self.stats["fixtures"]["unparseable_details"].append(
                    f"Row {row_idx}: Unparseable score string '{score_str}' for match {squad_a['name']} vs {squad_b['name']}"
                )
                return

        match_uuid = str(uuid.uuid5(uuid.NAMESPACE_DNS, f"match.{sport_obj['id']}.{gender}.league.{squad_a['id']}.{squad_b['id']}.{row_idx}"))
        match_record = {
            "id": match_uuid,
            "sport_id": sport_obj["id"],
            "team_a_id": squad_a["id"],
            "team_b_id": squad_b["id"],
            "gender": gender,
            "stage": "league",
            "level": "HS",
            "status": status,
            "round_info": "League Game" if not score_str else "Completed Game",
            "winner_team_id": winner_id,
            "is_draw": is_draw,
            "score_team_a": score_a,
            "score_team_b": score_b,
            "score_difference": score_diff,
            "score_summary": summary
        }

        self.stats["fixtures"].setdefault("created_list", []).append(match_record)
        if self.client:
            try:
                self.client.table("matches").insert(match_record).execute()
                self.stats["fixtures"]["created"] += 1
            except Exception as e:
                print(f"    Error inserting match: {e}")
        else:
            self.stats["fixtures"]["created"] += 1

    def resolve_squad(self, team_str: str, sport_id: str, gender: str) -> Optional[Dict[str, Any]]:
        for h_key, h_obj in self.houses_map.items():
            if h_key in team_str.lower():
                squad_label = "B" if (" b" in team_str.lower() or "(b)" in team_str.lower()) else ("A" if (" a" in team_str.lower() or "(a)" in team_str.lower()) else "Single")
                squad_key = f"{h_obj['id']}_{sport_id}_{gender}_{squad_label}"
                if squad_key in self.squads_map:
                    return self.squads_map[squad_key]

                for label in ["Single", "A", "B"]:
                    alt_key = f"{h_obj['id']}_{sport_id}_{gender}_{label}"
                    if alt_key in self.squads_map:
                        return self.squads_map[alt_key]
        return None

    def print_summary(self):
        print("\n================ SEEDING SUMMARY REPORT ================")
        print(f"  Houses:            {self.stats['houses']['created']} created/upserted")
        print(f"  Sports:            {self.stats['sports']['created']} created/upserted")
        print(f"  Tournament Groups: {self.stats['groups']['created']} created/upserted")
        print(f"  Squads:            {self.stats['squads']['created']} created/upserted")
        print(f"  Players:           {self.stats['players']['created']} created/upserted")
        print(f"  Fixtures:          {self.stats['fixtures']['created']} created")
        print(f"    - Unplayed fixtures (scheduled):   {self.stats['fixtures']['unplayed']}")
        print(f"    - Unparseable scores (logged):      {self.stats['fixtures']['unparseable']}")

        if self.stats["fixtures"]["unparseable_details"]:
            print("\n  Unparseable Fixtures Details:")
            for detail in self.stats["fixtures"]["unparseable_details"]:
                print(f"    * {detail}")
        print("========================================================")

    def run(self):
        file_loaded = self.load_workbook()
        self.seed_houses()
        self.seed_sports()
        self.ensure_default_squads()
        if file_loaded:
            self.parse_house_rosters()
            self.parse_fixtures_and_scores()
        self.print_summary()

    def generate_sql(self, output_path: str = "seed.sql"):
        print(f"\n=== Generating SQL Seed File: {output_path} ===")
        lines = [
            "-- Generated static seed.sql for DSS Sports Inter-House Meet",
            "-- Upsert-safe SQL file runnable directly in Supabase SQL editor or psql",
            "BEGIN;",
            ""
        ]

        # 1. Houses
        lines.append("-- 1. HOUSES")
        for h in self.houses_map.values():
            val_name = h['name'].replace("'", "''")
            val_code = h['short_code'].replace("'", "''")
            lines.append(
                f"INSERT INTO public.houses (id, name, color_hex, short_code) "
                f"VALUES ('{h['id']}', '{val_name}', '{h['color_hex']}', '{val_code}') "
                f"ON CONFLICT (name) DO UPDATE SET color_hex = EXCLUDED.color_hex, short_code = EXCLUDED.short_code;"
            )

        # 2. Sports
        lines.append("\n-- 2. SPORTS")
        for s in self.sports_map.values():
            val_name = s['name'].replace("'", "''")
            lines.append(
                f"INSERT INTO public.sports (id, name, type, level) "
                f"VALUES ('{s['id']}', '{val_name}', '{s['type']}', '{s['level']}') "
                f"ON CONFLICT (name) DO UPDATE SET type = EXCLUDED.type, level = EXCLUDED.level;"
            )

        # 3. Tournament Groups
        lines.append("\n-- 3. TOURNAMENT GROUPS")
        for g in self.groups_map.values():
            lines.append(
                f"INSERT INTO public.tournament_groups (id, sport_id, gender, format, point_win, point_draw, point_loss) "
                f"VALUES ('{g['id']}', '{g['sport_id']}', '{g['gender']}', '{g['format']}', {g['point_win']}, {g['point_draw']}, {g['point_loss']}) "
                f"ON CONFLICT (sport_id, gender) DO UPDATE SET format = EXCLUDED.format;"
            )

        # 4. Squads / Teams
        lines.append("\n-- 4. TEAMS / SQUADS")
        for sq in self.squads_map.values():
            val_name = sq['name'].replace("'", "''")
            pool_val = f"'{sq['pool']}'" if sq.get('pool') else "NULL"
            lines.append(
                f"INSERT INTO public.teams (id, name, house_id, gender, squad_label, pool, sport_id, level) "
                f"VALUES ('{sq['id']}', '{val_name}', '{sq['house_id']}', '{sq['gender']}', '{sq['squad_label']}', {pool_val}, '{sq['sport_id']}', '{sq['level']}') "
                f"ON CONFLICT (house_id, sport_id, gender, squad_label, level) DO UPDATE SET pool = EXCLUDED.pool, name = EXCLUDED.name;"
            )

        # 5. Players
        lines.append("\n-- 5. PLAYERS")
        for p in self.players_map.values():
            val_name = p['name'].replace("'", "''")
            roll_val = f"'{p['roll_number'].replace('\'', '\'\'')}'" if p.get('roll_number') else "NULL"
            grade_val = f"'{p['grade'].replace('\'', '\'\'')}'" if p.get('grade') else "NULL"
            sec_val = f"'{p['section'].replace('\'', '\'\'')}'" if p.get('section') else "NULL"

            lines.append(
                f"INSERT INTO public.players (id, name, team_id, roll_number, grade, section, gender, level) "
                f"VALUES ('{p['id']}', '{val_name}', '{p['team_id']}', {roll_num_sql(p.get('roll_number'))}, {str_sql(p.get('grade'))}, {str_sql(p.get('section'))}, '{p['gender']}', '{p['level']}') "
                f"ON CONFLICT (roll_number) DO UPDATE SET name = EXCLUDED.name, team_id = EXCLUDED.team_id, grade = EXCLUDED.grade, section = EXCLUDED.section;"
            )

        # 6. Matches
        lines.append("\n-- 6. MATCHES")
        created_matches = self.stats["fixtures"].get("created_list", [])
        for m in created_matches:
            team_a_val = f"'{m['team_a_id']}'" if m.get('team_a_id') else "NULL"
            team_b_val = f"'{m['team_b_id']}'" if m.get('team_b_id') else "NULL"
            winner_val = f"'{m['winner_team_id']}'" if m.get('winner_team_id') else "NULL"
            info_val = f"'{m['round_info'].replace('\'', '\'\'')}'" if m.get('round_info') else "''"
            summary_val = f"'{m['score_summary'].replace('\'', '\'\'')}'" if m.get('score_summary') else "''"
            is_draw_val = "TRUE" if m.get('is_draw') else "FALSE"

            lines.append(
                f"INSERT INTO public.matches (id, sport_id, team_a_id, team_b_id, gender, stage, level, status, round_info, winner_team_id, is_draw, score_team_a, score_team_b, score_difference, score_summary) "
                f"VALUES ('{m['id']}', '{m['sport_id']}', {team_a_val}, {team_b_val}, '{m['gender']}', '{m['stage']}', '{m['level']}', '{m['status']}', {info_val}, {winner_val}, {is_draw_val}, {m['score_team_a']}, {m['score_team_b']}, {m['score_difference']}, {summary_val}) "
                f"ON CONFLICT (id) DO UPDATE SET status = EXCLUDED.status, winner_team_id = EXCLUDED.winner_team_id, is_draw = EXCLUDED.is_draw, score_team_a = EXCLUDED.score_team_a, score_team_b = EXCLUDED.score_team_b, score_difference = EXCLUDED.score_difference, score_summary = EXCLUDED.score_summary;"
            )

        lines.append("\nCOMMIT;")

        with open(output_path, "w", encoding="utf-8") as f:
            f.write("\n".join(lines) + "\n")
        print(f"Successfully wrote {len(lines)} SQL lines to '{output_path}'")

def str_sql(val: Optional[str]) -> str:
    if val is None or val == "":
        return "NULL"
    escaped = str(val).replace("'", "''")
    return f"'{escaped}'"

def roll_num_sql(val: Optional[str]) -> str:
    if val is None or val == "":
        return "NULL"
    escaped = str(val).replace("'", "''")
    return f"'{escaped}'"

def main():
    parser = argparse.ArgumentParser(description="Seed DSS Sports Inter-House Meet data.")
    parser.add_argument("--file", default="seed-data/interhouse_meet.xlsx", help="Path to .xlsx file")
    parser.add_argument("--output-sql", action="store_true", help="Generate static SQL file instead of running API seeder")
    parser.add_argument("--sql-file", default="seed.sql", help="Output path for SQL seed file")
    args = parser.parse_args()

    supabase_client = None
    if not args.output_sql and SUPABASE_URL and SUPABASE_SERVICE_ROLE_KEY and "mock" not in SUPABASE_URL:
        try:
            from supabase import create_client
            supabase_client = create_client(SUPABASE_URL, SUPABASE_SERVICE_ROLE_KEY)
            print(f"Connected to Supabase at {SUPABASE_URL}")
        except Exception as e:
            print(f"Could not initialize Supabase client: {e}. Running in dry-run mode.")

    seeder = InterHouseSeeder(args.file, supabase_client)
    seeder.run()

    if args.output_sql:
        seeder.generate_sql(args.sql_file)

if __name__ == "__main__":
    main()
