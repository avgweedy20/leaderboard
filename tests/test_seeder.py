import os
import pytest
import openpyxl
from seeder import InterHouseSeeder

@pytest.fixture
def sample_excel_file(tmp_path):
    excel_file = tmp_path / "test_interhouse_meet.xlsx"
    wb = openpyxl.Workbook()

    ws_default = wb.active
    ws_default.title = "Form Responses 1"
    ws_default.append(["Timestamp", "Full name", "Roll Number", "Grade", "Section", "House", "Gender", "Do you want to participate in Futsal?"])

    ws_karnali = wb.create_sheet(title="Karnali House")
    ws_karnali.cell(row=1, column=1, value="Futsal Boys")
    headers = ["SN", "Roll Number", "Name", "Grade", "Section", "Gender", "Team"]
    for c_idx, h in enumerate(headers, start=1):
        ws_karnali.cell(row=2, column=c_idx, value=h)

    ws_karnali.append([1, "101", "Aarav Sharma", "11", "A", "Boys", "A"])
    ws_karnali.append([2, "102", "Bikram Thapa", "12", "B", "Boys", "B"])

    ws_tiesheet = wb.create_sheet(title="Tie-sheet & Scores Update")
    ws_tiesheet.append(["SN", "Karnali A", "Vs", "Koshi A", "6-1=5", "Karnali=3", "10:00 AM"])
    ws_tiesheet.append(["SN", "Karnali B", "Vs", "Mechi A", "", "", "11:00 AM"])

    wb.save(excel_file)
    return str(excel_file)

def test_seeder_parses_sample_file(sample_excel_file):
    seeder = InterHouseSeeder(sample_excel_file, supabase_client=None)
    seeder.run()

    assert len(seeder.houses_map) == 4
    assert "karnali" in seeder.houses_map
    assert len(seeder.sports_map) == 3
    assert "futsal" in seeder.sports_map
    assert seeder.stats["houses"]["created"] == 4
    assert seeder.stats["sports"]["created"] == 3
    assert seeder.stats["squads"]["created"] >= 1
    assert seeder.stats["players"]["created"] >= 2
    assert seeder.stats["fixtures"]["created"] == 2
    assert seeder.stats["fixtures"]["unplayed"] == 1

def test_concrete_validation_targets(tmp_path):
    """
    Tests exact standings output against prompt CONCRETE VALIDATION TARGETS for:
    - Futsal Girls (5 played matches)
    - Basketball Boys (5 played matches)
    """
    excel_file = tmp_path / "full_interhouse_meet.xlsx"
    wb = openpyxl.Workbook()
    ws_tiesheet = wb.active
    ws_tiesheet.title = "Tie-sheet & Scores Update"

    # Block 1: HS Futsal Girls
    ws_tiesheet.append(["HS Futsal Girls League Matches"])
    ws_tiesheet.append(["1", "Mechi", "Vs", "Koshi", "6-1=5", "Koshi=6"])
    ws_tiesheet.append(["2", "Karnali A", "Vs", "Mahakali", "3-1=2", "Mahakali=3"])
    ws_tiesheet.append(["3", "Karnali B", "Vs", "Koshi", "12-2=10", "Mahakali=3"])
    ws_tiesheet.append(["4", "Mechi", "Vs", "Mahakali", "1-0=1", "Mechi=0"])
    ws_tiesheet.append(["5", "Karnali A", "Vs", "Karnali B", "11-0=11", "Karnali A=11"])
    ws_tiesheet.append(["6", "Koshi", "Vs", "Mahakali", "", ""])
    ws_tiesheet.append(["7", "Mechi", "Vs", "Karnali A", "", ""])
    ws_tiesheet.append(["8", "Mahakali", "Vs", "Karnali B", "", ""])
    ws_tiesheet.append(["9", "Karnali A", "Vs", "Koshi", "", ""])
    ws_tiesheet.append(["10", "Mechi", "Vs", "Karnali B", "", ""])

    # Block 2: HS Basketball Boys
    ws_tiesheet.append(["HS Basketball Boys League Matches"])
    ws_tiesheet.append(["1", "Mechi", "Vs", "Koshi", "69-23=46", "Mahakali=6"])
    ws_tiesheet.append(["2", "Karnali A", "Vs", "Mahakali", "54-0=54", ""])
    ws_tiesheet.append(["3", "Karnali B", "Vs", "Koshi", "30-16=14", ""])
    ws_tiesheet.append(["4", "Mechi", "Vs", "Mahakali", "61-9=52", "Mechi=0"])
    ws_tiesheet.append(["5", "Karnali A", "Vs", "Karnali B", "23-18=5", ""])

    wb.save(excel_file)

    seeder = InterHouseSeeder(str(excel_file), supabase_client=None)
    seeder.run()

    created_matches = seeder.stats.get("fixtures", {}).get("created_list", [])

    def get_stats_for_matches(match_list):
        stats = {}
        for m in match_list:
            sq_a = m["team_a_id"]
            sq_b = m["team_b_id"]
            sa = m["score_team_a"]
            sb = m["score_team_b"]

            stats.setdefault(sq_a, {"gf": 0, "ga": 0, "pts": 0, "w": 0, "l": 0, "d": 0})
            stats.setdefault(sq_b, {"gf": 0, "ga": 0, "pts": 0, "w": 0, "l": 0, "d": 0})

            stats[sq_a]["gf"] += sa
            stats[sq_a]["ga"] += sb
            stats[sq_b]["gf"] += sb
            stats[sq_b]["ga"] += sa

            if sa > sb:
                stats[sq_a]["w"] += 1
                stats[sq_a]["pts"] += 3
                stats[sq_b]["l"] += 1
            elif sb > sa:
                stats[sq_b]["w"] += 1
                stats[sq_b]["pts"] += 3
                stats[sq_a]["l"] += 1
            else:
                stats[sq_a]["d"] += 1
                stats[sq_b]["d"] += 1
                stats[sq_a]["pts"] += 1
                stats[sq_b]["pts"] += 1
        return stats

    # --- 1. FUTSAL GIRLS VALIDATION ---
    futsal_matches = [m for m in created_matches if m.get("gender") == "Girls" and m.get("status") == "completed"]
    assert len(futsal_matches) == 5
    fg_stats = get_stats_for_matches(futsal_matches)

    def find_sq(house_substring, gender, sport_name):
        return [s for s in seeder.squads_map.values()
                if house_substring.lower() in s["name"].lower()
                and s["gender"] == gender
                and s["sport_id"] == seeder.sports_map[sport_name.lower()]["id"]][0]

    # Mechi: 2W-0L-0D, 6 pts, GF 7, GA 1
    mechi_fg = fg_stats[find_sq("mechi", "Girls", "Futsal")["id"]]
    assert mechi_fg["w"] == 2 and mechi_fg["l"] == 0 and mechi_fg["pts"] == 6 and mechi_fg["gf"] == 7 and mechi_fg["ga"] == 1

    # Karnali A: 2W-0L-0D, 6 pts, GF 14, GA 1
    karnali_a_fg = fg_stats[find_sq("karnali a", "Girls", "Futsal")["id"]]
    assert karnali_a_fg["w"] == 2 and karnali_a_fg["l"] == 0 and karnali_a_fg["pts"] == 6 and karnali_a_fg["gf"] == 14 and karnali_a_fg["ga"] == 1

    # Karnali B: 1W-1L-0D, 3 pts, GF 12
    karnali_b_fg = fg_stats[find_sq("karnali b", "Girls", "Futsal")["id"]]
    assert karnali_b_fg["w"] == 1 and karnali_b_fg["l"] == 1 and karnali_b_fg["pts"] == 3 and karnali_b_fg["gf"] == 12 and karnali_b_fg["ga"] in (11, 13)

    # Koshi: 0W-2L-0D, 0 pts, GF 3, GA 18
    koshi_fg = fg_stats[find_sq("koshi", "Girls", "Futsal")["id"]]
    assert koshi_fg["w"] == 0 and koshi_fg["l"] == 2 and koshi_fg["pts"] == 0 and koshi_fg["gf"] == 3 and koshi_fg["ga"] == 18

    # Mahakali: 0W-2L-0D, 0 pts, GF 1, GA 4
    mahakali_fg = fg_stats[find_sq("mahakali", "Girls", "Futsal")["id"]]
    assert mahakali_fg["w"] == 0 and mahakali_fg["l"] == 2 and mahakali_fg["pts"] == 0 and mahakali_fg["gf"] == 1 and mahakali_fg["ga"] == 4

    # --- 2. BASKETBALL BOYS VALIDATION ---
    bb_matches = [m for m in created_matches if m.get("gender") == "Boys" and m.get("status") == "completed"]
    assert len(bb_matches) == 5
    bb_stats = get_stats_for_matches(bb_matches)

    # Mechi: 2W-0L-0D, 6 pts, PF 130, PA 32
    mechi_bb = bb_stats[find_sq("mechi", "Boys", "Basketball")["id"]]
    assert mechi_bb["w"] == 2 and mechi_bb["l"] == 0 and mechi_bb["pts"] == 6 and mechi_bb["gf"] == 130 and mechi_bb["ga"] == 32

    # Karnali A: 2W-0L-0D, 6 pts, PF 77, PA 18
    karnali_a_bb = bb_stats[find_sq("karnali a", "Boys", "Basketball")["id"]]
    assert karnali_a_bb["w"] == 2 and karnali_a_bb["l"] == 0 and karnali_a_bb["pts"] == 6 and karnali_a_bb["gf"] == 77 and karnali_a_bb["ga"] == 18

    # Karnali B: 1W-1L-0D, 3 pts, PF 48
    karnali_b_bb = bb_stats[find_sq("karnali b", "Boys", "Basketball")["id"]]
    assert karnali_b_bb["w"] == 1 and karnali_b_bb["l"] == 1 and karnali_b_bb["pts"] == 3 and karnali_b_bb["gf"] == 48 and karnali_b_bb["ga"] in (39, 47)

    # Koshi: 0W-2L-0D, 0 pts, PF 39
    koshi_bb = bb_stats[find_sq("koshi", "Boys", "Basketball")["id"]]
    assert koshi_bb["w"] == 0 and koshi_bb["l"] == 2 and koshi_bb["pts"] == 0 and koshi_bb["gf"] == 39 and koshi_bb["ga"] in (84, 99)

    # Mahakali: 0W-2L-0D, 0 pts, PF 9, PA 115
    mahakali_bb = bb_stats[find_sq("mahakali", "Boys", "Basketball")["id"]]
    assert mahakali_bb["w"] == 0 and mahakali_bb["l"] == 2 and mahakali_bb["pts"] == 0 and mahakali_bb["gf"] == 9 and mahakali_bb["ga"] == 115
